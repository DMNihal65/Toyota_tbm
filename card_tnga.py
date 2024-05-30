from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
import smtplib
from datetime import datetime, timedelta, date
import requests
from prettytable import PrettyTable
from openpyxl import Workbook
import json
from openpyxl.styles import Font


today = datetime.now()
from_date = today-timedelta(days=30)
to_date = today+timedelta(days=1)

f_y = from_date.year
f_m = from_date.month
f_d = from_date.day

t_y = to_date.year
t_m = to_date.month
t_d = to_date.day

if today.isoweekday() == 1:
    yesterday = today-timedelta(days=2)
else:
    yesterday = today-timedelta(days=1)


y_weekd = yesterday.isoweekday()
y_y = yesterday.year
y_m = yesterday.month
y_d = yesterday.day
y_weekNo = (y_d//7.1)+1
yesterday_date = f"{y_d}-{y_m}-{y_y}"
# print(yesterday_date)


weekd = datetime.now().isoweekday()
y = datetime.now().year
m = datetime.now().month
d = datetime.now().day
# to fetch week no.
weekNo = (d//7.1)+1

# generate the url address
urlForCardRaised = "http://10.82.126.73:5071/card/find/fromDate/" + \
    str(f_y)+"-"+str(f_m)+"-"+str(f_d)+"/toDate/" + \
    str(t_y)+"-"+str(t_m)+"-"+str(t_d)
urlForCardRaised_4_yesterday = "http://10.82.126.73:5071/card/find/fromDate/" + \
    str(y_y)+"-"+str(y_m)+"-"+str(y_d)+"/toDate/"+str(y)+"-"+str(m)+"-"+str(d)
print("URL-last 30 days", urlForCardRaised)
print("URL-yesterday", urlForCardRaised_4_yesterday)


# fetching the data using api
response = requests.get(urlForCardRaised, auth=('user', 'pass'))                  
print(response.status_code)
card = response.json()
# print(card["cards"])

response2 = requests.get(urlForCardRaised_4_yesterday, auth=('user', 'pass'))
print(response2.status_code)
card_y = response2.json()
#####################################################################################################
wb = Workbook()

wsS = wb.create_sheet("Summary", 0)
wsB = wb.create_sheet("Block", 1)
wsC = wb.create_sheet("Crank", 2)
wsH = wb.create_sheet("Head", 3)


wsS.column_dimensions['A'].width = 12
wsS.column_dimensions['B'].width = 12
wsS.column_dimensions['C'].width = 12
wsS.column_dimensions['D'].width = 12
wsS.column_dimensions['E'].width = 12
wsB.column_dimensions['A'].width = 12
wsB.column_dimensions['B'].width = 12
wsB.column_dimensions['C'].width = 12
wsB.column_dimensions['D'].width = 12
wsB.column_dimensions['E'].width = 12
wsC.column_dimensions['A'].width = 12
wsC.column_dimensions['B'].width = 12
wsC.column_dimensions['C'].width = 12
wsC.column_dimensions['D'].width = 12
wsC.column_dimensions['E'].width = 12
wsH.column_dimensions['A'].width = 12
wsH.column_dimensions['B'].width = 12
wsH.column_dimensions['C'].width = 12
wsH.column_dimensions['D'].width = 12
wsH.column_dimensions['E'].width = 12

if card["success"] == True:
    card_list = card["cards"]  # last 30 cards
if card_y["success"] == True:  # yesterdays cards
    card_list_y = card_y["cards"]

if card["success"] == True:
    # ================================Block Line=========================================================
    r = 3
    card_message = "Daily smile card raise status in TNGA is attached."
    for entry in card_list:
        if entry["line"] == "Block" and entry["status"] != "complete":
            wsB["A1"] = "Summary of card raised in TNGA BLOCK LINE (last 30 days)"
            wsB['A1'].font = Font(bold=True)
            wsB["A2"] = "OP No."
            wsB['A2'].font = Font(bold=True)
            wsB["B2"] = "Status"
            wsB['B2'].font = Font(bold=True)
            wsB["C2"] = "Abnormality"
            wsB['C2'].font = Font(bold=True)
            wsB["D2"] = "Card"
            wsB['D2'].font = Font(bold=True)
            wsB["E2"] = "Pending Days"
            wsB["E2"].font = Font(bold=True)
            # print(entry["line"],",",entry["status"])

            opNo = wsB.cell(row=r, column=1)
            opNo.value = entry["processNo"]
            status = wsB.cell(row=r, column=2)
            status.value = entry["status"]
            abnormality = wsB.cell(row=r, column=3)
            abnormality.value = entry["abnormality"]
            cardType = wsB.cell(row=r, column=4)
            cardType.value = entry["cardType"]
            pending_days = wsB.cell(row=r, column=5)
            dateCreated = date(int(entry["createdAt"][0:4]), int(
                entry["createdAt"][5:7]), int(entry["createdAt"][8:10]))
            pending_days.value = (today.date()-dateCreated).days
            r = r+1
    # ================================Crank Line=========================================================
    r = 3
    for entry in card_list:
        if entry["line"] == "Crank" and entry["status"] != "complete":
            wsC["A1"] = "Summary of card raised in TNGA CRANK LINE (last 30 days)"
            wsC['A1'].font = Font(bold=True)
            wsC["A2"] = "OP No."
            wsC['A2'].font = Font(bold=True)
            wsC["B2"] = "Status"
            wsC['B2'].font = Font(bold=True)
            wsC["C2"] = "Abnormality"
            wsC['C2'].font = Font(bold=True)
            wsC["D2"] = "Card"
            wsC['D2'].font = Font(bold=True)
            wsC["E2"] = "Pending Days"
            wsC['E2'].font = Font(bold=True)
            # print(entry["line"],",",entry["status"])
            opNo = wsC.cell(row=r, column=1)
            opNo.value = entry["processNo"]
            status = wsC.cell(row=r, column=2)
            status.value = entry["status"]
            abnormality = wsC.cell(row=r, column=3)
            abnormality.value = entry["abnormality"]
            cardType = wsC.cell(row=r, column=4)
            cardType.value = entry["cardType"]
            pending_days = wsC.cell(row=r, column=5)
            dateCreated = date(int(entry["createdAt"][0:4]), int(
                entry["createdAt"][5:7]), int(entry["createdAt"][8:10]))
            pending_days.value = (today.date()-dateCreated).days
            r = r+1
    # ================================HEAD Line=========================================================
    r = 3
    for entry in card_list:
        if entry["line"] == "Head" and entry["status"] != "complete":
            wsH["A1"] = "Summary of card raised in TNGA HEAD LINE (last 30 days)"
            wsH['A1'].font = Font(bold=True)
            wsH["A2"] = "OP No."
            wsH['A2'].font = Font(bold=True)
            wsH["B2"] = "Status"
            wsH['B2'].font = Font(bold=True)
            wsH["C2"] = "Abnormality"
            wsH['C2'].font = Font(bold=True)
            wsH["D2"] = "Card"
            wsH['D2'].font = Font(bold=True)
            wsH["E2"] = "Pending Days"
            wsH['E2'].font = Font(bold=True)
            # print(entry["line"],",",entry["status"])
            opNo = wsH.cell(row=r, column=1)
            opNo.value = entry["processNo"]
            status = wsH.cell(row=r, column=2)
            status.value = entry["status"]
            abnormality = wsH.cell(row=r, column=3)
            abnormality.value = entry["abnormality"]
            cardType = wsH.cell(row=r, column=4)
            cardType.value = entry["cardType"]
            pending_days = wsH.cell(row=r, column=5)
            dateCreated = date(int(entry["createdAt"][0:4]), int(
                entry["createdAt"][5:7]), int(entry["createdAt"][8:10]))
            pending_days.value = (today.date()-dateCreated).days
            r = r+1
    # ================================SUMMARY SHEET-Last 30 days=========================================================
    wsS["A1"] = "Summary of card raised in TNGA (last 30 days)"
    wsS['A1'].font = Font(bold=True)
    wsS["A2"] = "LINE"
    wsS['A2'].font = Font(bold=True)
    wsS["B2"] = "TOTAL"
    wsS['B2'].font = Font(bold=True)
    wsS["C2"] = "COMPLETE"
    wsS['C2'].font = Font(bold=True)
    wsS["D2"] = "INPROGRESS"
    wsS['D2'].font = Font(bold=True)
    wsS["E2"] = "PENDING"
    wsS['E2'].font = Font(bold=True)
    wsS["A3"] = "BLOCK"
    wsS['A3'].font = Font(bold=True)
    wsS["A4"] = "CRANK"
    wsS['A4'].font = Font(bold=True)
    wsS["A5"] = "HEAD"
    wsS['A5'].font = Font(bold=True)

    completeB = 0
    completeC = 0
    completeH = 0
    inprogressB = 0
    inprogressC = 0
    inprogressH = 0
    pendingB = 0
    pendingC = 0
    pendingH = 0
    for entry in card_list:
        if entry["line"] == "Head":
            if entry["status"] == "complete":
                completeH = completeH+1
            if entry["status"] == "inprogress":
                inprogressH = inprogressH+1
            if entry["status"] == "pending":
                pendingH = pendingH+1
        if entry["line"] == "Crank":
            if entry["status"] == "complete":
                completeC = completeC+1
            if entry["status"] == "inprogress":
                inprogressC = inprogressC+1
            if entry["status"] == "pending":
                pendingC = pendingC+1
        if entry["line"] == "Block":
            if entry["status"] == "complete":
                completeB = completeB+1
            if entry["status"] == "inprogress":
                inprogressB = inprogressB+1
            if entry["status"] == "pending":
                pendingB = pendingB+1

    # ===========summary of BLOCK===================
    total_B = wsS.cell(row=3, column=2)
    total_B.value = completeB+inprogressB+pendingB
    complete_B = wsS.cell(row=3, column=3)
    complete_B.value = completeB
    inprogress_B = wsS.cell(row=3, column=4)
    inprogress_B.value = inprogressB
    pending_B = wsS.cell(row=3, column=5)
    pending_B.value = pendingB

    # ===========summary of CRANK===================
    total_C = wsS.cell(row=4, column=2)
    total_C.value = completeC+inprogressC+pendingC
    complete_C = wsS.cell(row=4, column=3)
    complete_C.value = completeC
    inprogress_C = wsS.cell(row=4, column=4)
    inprogress_C.value = inprogressC
    pending_C = wsS.cell(row=4, column=5)
    pending_C.value = pendingC

    # ===========summary of HEAD===================
    total_H = wsS.cell(row=5, column=2)
    total_H.value = completeH+inprogressH+pendingH
    complete_H = wsS.cell(row=5, column=3)
    complete_H.value = completeH
    inprogress_H = wsS.cell(row=5, column=4)
    inprogress_H.value = inprogressH
    pending_H = wsS.cell(row=5, column=5)
    pending_H.value = pendingH
    # ================================SUMMARY SHEET-YESTERDAY=========================================================
    if card_y["success"] == True:
        wsS["A11"] = "Summary of card raised in TNGA (Yesterday)"
        wsS['A11'].font = Font(bold=True)
        wsS["A12"] = "LINE"
        wsS['A12'].font = Font(bold=True)
        wsS["B12"] = "TOTAL"
        wsS['B12'].font = Font(bold=True)
        wsS["C12"] = "COMPLETE"
        wsS['C12'].font = Font(bold=True)
        wsS["D12"] = "INPROGRESS"
        wsS['D12'].font = Font(bold=True)
        wsS["E12"] = "PENDING"
        wsS['E12'].font = Font(bold=True)
        wsS["A13"] = "BLOCK"
        wsS['A13'].font = Font(bold=True)
        wsS["A14"] = "CRANK"
        wsS['A14'].font = Font(bold=True)
        wsS["A15"] = "HEAD"
        wsS['A15'].font = Font(bold=True)

        completeB = 0
        completeC = 0
        completeH = 0
        inprogressB = 0
        inprogressC = 0
        inprogressH = 0
        pendingB = 0
        pendingC = 0
        pendingH = 0
        for entry in card_list_y:
            if entry["line"] == "Head":
                if entry["status"] == "complete":
                    completeH = completeH+1
                if entry["status"] == "inprogress":
                    inprogressH = inprogressH+1
                if entry["status"] == "pending":
                    pendingH = pendingH+1
            if entry["line"] == "Crank":
                if entry["status"] == "complete":
                    completeC = completeC+1
                if entry["status"] == "inprogress":
                    inprogressC = inprogressC+1
                if entry["status"] == "pending":
                    pendingC = pendingC+1
            if entry["line"] == "Block":
                if entry["status"] == "complete":
                    completeB = completeB+1
                if entry["status"] == "inprogress":
                    inprogressB = inprogressB+1
                if entry["status"] == "pending":
                    pendingB = pendingB+1

        # ===========summary of BLOCK===================
        total_B = wsS.cell(row=13, column=2)
        total_B.value = completeB+inprogressB+pendingB
        complete_B = wsS.cell(row=13, column=3)
        complete_B.value = completeB
        inprogress_B = wsS.cell(row=13, column=4)
        inprogress_B.value = inprogressB
        pending_B = wsS.cell(row=13, column=5)
        pending_B.value = pendingB

        # ===========summary of CRANK===================
        total_C = wsS.cell(row=14, column=2)
        total_C.value = completeC+inprogressC+pendingC
        complete_C = wsS.cell(row=14, column=3)
        complete_C.value = completeC
        inprogress_C = wsS.cell(row=14, column=4)
        inprogress_C.value = inprogressC
        pending_C = wsS.cell(row=14, column=5)
        pending_C.value = pendingC

        # ===========summary of HEAD===================
        total_H = wsS.cell(row=15, column=2)
        total_H.value = completeH+inprogressH+pendingH
        complete_H = wsS.cell(row=15, column=3)
        complete_H.value = completeH
        inprogress_H = wsS.cell(row=15, column=4)
        inprogress_H.value = inprogressH
        pending_H = wsS.cell(row=15, column=5)
        pending_H.value = pendingH
        # ====================================================================================================

    wb.save("card_status_tnga.xlsx")
else:
    card_message = "No TBM/OM card raised in last 30 days."
# ==============================excel part finished=========================================
# ==============================mail summarry part starts===================================

today = datetime.now()
if today.isoweekday() == 1:
    yesterday = today-timedelta(days=2)
else:
    yesterday = today-timedelta(days=1)


y_weekd = yesterday.isoweekday()
y_y = yesterday.year
y_m = yesterday.month
y_d = yesterday.day
y_weekNo = (y_d//7.1)+1

weekd = datetime.now().isoweekday()
y = datetime.now().year
m = datetime.now().month
d = datetime.now().day
# to fetch week no.
weekNo = (d//7.1)+1

# generate the url address
# print("Year:",y,"Month:",m,"Day:",d,"WeekDay:",weekd,"Week No:",weekNo)
url4mcDetails = "http://10.82.126.73:5071/head/headMachineList?d=" + \
    str(y_weekd)+"&y="+str(y_y)+"&w="+str(y_weekNo)+"&m="+str(y_m)+"&pS=P"
url4dailyStatus = "http://10.82.126.73:5071/dailyStatus?entryFor=" + \
    str(y_y)+"-"+str(y_m)+"-"+str(y_d)+"&pS=P"
url4mcDetails_OM = "http://10.82.126.73:5071/head/headMachineList?d=" + \
    str(y_weekd)+"&y="+str(y_y)+"&w="+str(y_weekNo)+"&m="+str(y_m)+"&pS=S"
url4dailyStatus_OM = "http://10.82.126.73:5071/dailyStatus?entryFor=" + \
    str(y_y)+"-"+str(y_m)+"-"+str(y_d)+"&pS=S"
print(url4mcDetails)
print(url4dailyStatus)
print(url4mcDetails_OM)
print(url4dailyStatus_OM)

# fetching the data using api
response1 = requests.get(url4mcDetails, auth=('user', 'pass'))
response2 = requests.get(url4dailyStatus, auth=('user', 'pass'))
mcDetails = response1.json()
dailyStatus = response2.json()

response3 = requests.get(url4mcDetails_OM, auth=('user', 'pass'))
response4 = requests.get(url4dailyStatus_OM, auth=('user', 'pass'))
mcDetails_OM = response3.json()
dailyStatus_OM = response4.json()

# ================================mail summary for maintenance===============================================================
line_list_mnt = []
tabular_fields = ["Line", "Total", "OK", "NG", "Pending"]
tabular_table = PrettyTable()
tabular_table.field_names = tabular_fields


for i in mcDetails["machineData"]:
    # m=mcData["machineData"]
    # print("from-mcData:",i["line"])
    line_name_mcData = i["line"]
    # print(i["counts"])
    nn = i["counts"]
    total_points = sum(nn.values())
    # print("Total:",sum(nn.values()))
    ok_p = 0
    ng_p = 0

    for j in dailyStatus["sortedDailyStatus"]:
        line_name_statusData = j["line"]
        if line_name_statusData == line_name_mcData:
            # ok_p=0
            # ng_p=0
            # print("from-sortedDetails:",j["line"])

            for jj in j["processes"]:
                # print(jj["result"],"\n")
                ok_p = ok_p + jj["result"].get("OK")
                ng_p = ng_p + jj["result"].get("NG")
            # print(j,"\n\n")
            # print("OK:",ok_p)
            # print("NG",ng_p)
            # print("\n")
        else:
            pass
    tabular_table.add_row(
        [line_name_mcData, total_points, ok_p, ng_p, total_points-ok_p-ng_p])
    line_list_mnt.append(line_name_mcData)
    ok_p = 0
    ng_p = 0

total_line_list_mnt = ["Assembly (Block Sub-assembly)", "Assembly (Head Sub-assembly)", "Assembly (MK-1)",
                       "Assembly (MK-2)", "Assembly (Piston Sub-assembly)", "Assembly (Test Bench)", "Block", "Crank", "Head"]
for line_name_mnt in total_line_list_mnt:
    if line_name_mnt not in line_list_mnt:
        tabular_table.add_row([line_name_mnt, "-", "-", "-", "-"])
# ================================mail summary for production================================================================
line_list_prod = []
tabular_fields_OM = ["Line", "Total", "OK", "NG", "Pending"]
tabular_table_OM = PrettyTable()
tabular_table_OM.field_names = tabular_fields_OM


for i in mcDetails_OM["machineData"]:
    # m=mcData["machineData"]
    # print("from-mcData:",i["line"])
    line_name_mcData = i["line"]
    # print(i["counts"])
    nn = i["counts"]
    total_points = sum(nn.values())
    # print("Total:",sum(nn.values()))
    ok_p = 0
    ng_p = 0

    for j in dailyStatus_OM["sortedDailyStatus"]:
        line_name_statusData = j["line"]
        if line_name_statusData == line_name_mcData:
            # ok_p=0
            # ng_p=0
            # print("from-sortedDetails:",j["line"])

            for jj in j["processes"]:
                # print(jj["result"],"\n")
                ok_p = ok_p + jj["result"].get("OK")
                ng_p = ng_p + jj["result"].get("NG")
            # print(j,"\n\n")
            # print("OK:",ok_p)
            # print("NG",ng_p)
            # print("\n")
        else:
            pass
    tabular_table_OM.add_row(
        [line_name_mcData, total_points, ok_p, ng_p, total_points-ok_p-ng_p])
    line_list_prod.append(line_name_mcData)
    ok_p = 0
    ng_p = 0

total_line_list_prod = ["Assembly (Block Sub-assembly)", "Assembly (Head Sub-assembly)", "Assembly (MK-1)",
                        "Assembly (MK-2)", "Assembly (Piston Sub-assembly)", "Assembly (Test Bench)", "Block", "Crank", "Head"]
for line_name_prod in total_line_list_prod:
    if line_name_prod not in line_list_prod:
        tabular_table_OM.add_row([line_name_prod, "-", "-", "-", "-"])
# ==============================mail summarry part ends=====================================

# ==============================table for mail-start========================================
print("OM\n", tabular_table_OM)
print("SM\n", tabular_table)
# ==============================table for mail-end========================================

# ==============================TO SEND EMAIL==TNGA MNT=============================================


# Connection with the server
server = smtplib.SMTP(host="smtp.office365.com", port=587)
server.starttls()
server.login("pankaj.jogi@tiei.toyota-industries.com", "WB4psCxK")

# Creation of the MIMEMultipart Object
message = MIMEMultipart()
# =================================================================================================
family = [    
    "naveen.kp@tiei.toyota-industries.com",
    "yousuf@tiei.toyota-industries.com",
    # "pankaj.jogi@tiei.toyota-industries.com",
    "puranik.kv@tiei.toyota-industries.com",
    "shridharbhat@tiei.toyota-industries.com",
    "srharidas@tiei.toyota-industries.com",
    # "lohith.s@tiei.toyota-industries.com",
    "manjunatha.ramadas@tiei.toyota-industries.com",
    "nagendra.ks@tiei.toyota-industries.com",
    # "kirankumar.r@tiei.toyota-industries.com",
    # "vijay.r@tiei.toyota-industries.com",
    # "vinothkumar.r@tiei.toyota-industries.com",
    # "kailash.kp@tiei.toyota-industries.com",
    "sundeep@tiei.toyota-industries.com",
    "sivasreekar.reddy@tiei.toyota-industries.com",
    # "sanketh@tiei.toyota-industries.com",
    # "kirana@tiei.toyota-industries.com",
    # "manu.b@tiei.toyota-industries.com",
    # "sandeep.a@tiei.toyota-industries.com",
    "marimuthu.k@tiei.toyota-industries.com"
 
   
]

# msg['To'] =', '.join(family)
# msg =', '.join(family)
# print(msg)
# =================================================================================================


# Setup of MIMEMultipart Object Header
message['From'] = "pankaj.jogi@tiei.toyota-industries.com"
# message['To'] = "pankaj.jogi@tiei.toyota-industries.com"
message['To'] = ', '.join(family)
message['Subject'] = "TNGA-MNT Daily TBM card status"

# Creation of a MIMEText Part
# textPart1 = MIMEText(f"Good Morning\nSUMMARY OF PREVIOUS DAY ({y_d}-{y_m}-{y_y}) TBM CHECK ACTIVITY\n\nA.PRODUCTION\n\nLine\tTotal\tOK\tNG\tPending\nBlock\t{total_block_OM}\t{ok_Block_OM}\t{ng_Block_OM}\t{total_block_OM-ok_Block_OM-ng_Block_OM}\nCrank\t{total_crank_OM}\t{ok_Crank_OM}\t{ng_Crank_OM}\t{total_crank_OM-ok_Crank_OM-ng_Crank_OM}\nHead\t{total_head_OM}\t{ok_Head_OM}\t{ng_Head_OM}\t{total_head_OM-ok_Head_OM-ng_Head_OM}\n\nB.MAINTENANCE\n\nLine\tTotal\tOK\tNG\tPending\nBlock\t{total_block}\t{ok_Block}\t{ng_Block}\t{total_block-ok_Block-ng_Block}\nCrank\t{total_crank}\t{ok_Crank}\t{ng_Crank}\t{total_crank-ok_Crank-ng_Crank}\nHead\t{total_head}\t{ok_Head}\t{ng_Head}\t{total_head-ok_Head-ng_Head}\n\nDaily TBM card raise status is attached.\nThank You\n\nMNT-MES", 'plain')
my_message = tabular_table_OM.get_html_string()
my_message2 = tabular_table.get_html_string()
html = """\
<html>
    <head>
    <style>
        table, th, td {
            border: 1px solid black;
            border-collapse: collapse;
        }
        th, td {
            padding: 5px;
            text-align: left;    
        }    
    </style>
    </head>
<body>
<p>
Good Morning<br> 
SUMMARY OF PREVIOUS DAY <b>(%s)</b> TBM CHECK ACTIVITY<br> 
</p>
<p>    
<b>TNGA-MAINTENANCE</b>
    <br>
    %s
</p>
<p>
%s<br>
Thank You
</p>
<p>
MNT-MES
</p>
</body>
</html>
""" % (yesterday_date, my_message2, card_message)

textPart1 = MIMEText(html, 'html')

# Creation of a MIMEApplication Part

if card["success"] == True:
    filename = "card_status_tnga.xlsx"
    filePart = MIMEApplication(open(filename, "rb").read(), Name=filename)
    filePart["Content-Disposition"] = 'attachment; filename="%s' % filename

    # Parts attachment
    message.attach(filePart)
message.attach(textPart1)

# Send Email and close connection
server.send_message(message)
server.quit()

# ==================================EMAIL PART END===============================================


# ==============================TO SEND EMAIL==TNGA PROD=============================================


# Connection with the server
server = smtplib.SMTP(host="smtp.office365.com", port=587)
server.starttls()
server.login("pankaj.jogi@tiei.toyota-industries.com", "WB4psCxK")

# Creation of the MIMEMultipart Object
message = MIMEMultipart()

# =================================================================================================
family = [

    "raghavendra.singh@tiei.toyota-industries.com",
    "yousuf@tiei.toyota-industries.com",
    "anjineya.b@tiei.toyota-industries.com",
    "gurunath.mn@tiei.toyota-industries.com",
    "raviverma.n@tiei.toyota-industries.com",
    "nitish.cr@tiei.toyota-industries.com",
    "mouneesh.m@tiei.toyota-industries.com",
    "sharath.k@tiei.toyota-industries.com",
    "vishwapoorna.rao@tiei.toyota-industries.com",
    "madhusudhan.h@tiei.toyota-industries.com",
    "srinath.waiker@tiei.toyota-industries.com",
    "nandeesha.n@tiei.toyota-industries.com",
   
]



# msg['To'] =', '.join(family)
# msg =', '.join(family)
# print(msg)
# =================================================================================================



# Setup of MIMEMultipart Object Header
message['From'] = "pankaj.jogi@tiei.toyota-industries.com"
# message['To'] = "pankaj.jogi@tiei.toyota-industries.com"
message['To'] = ', '.join(family)
message['Subject'] = "TNGA-PROD Daily OM card status"



# Creation of a MIMEText Part
# textPart1 = MIMEText(f"Good Morning\nSUMMARY OF PREVIOUS DAY ({y_d}-{y_m}-{y_y}) OM CHECK ACTIVITY\n\nA.PRODUCTION\n\nLine\tTotal\tOK\tNG\tPending\nBlock\t{total_block_OM}\t{ok_Block_OM}\t{ng_Block_OM}\t{total_block_OM-ok_Block_OM-ng_Block_OM}\nCrank\t{total_crank_OM}\t{ok_Crank_OM}\t{ng_Crank_OM}\t{total_crank_OM-ok_Crank_OM-ng_Crank_OM}\nHead\t{total_head_OM}\t{ok_Head_OM}\t{ng_Head_OM}\t{total_head_OM-ok_Head_OM-ng_Head_OM}\n\nB.MAINTENANCE\n\nLine\tTotal\tOK\tNG\tPending\nBlock\t{total_block}\t{ok_Block}\t{ng_Block}\t{total_block-ok_Block-ng_Block}\nCrank\t{total_crank}\t{ok_Crank}\t{ng_Crank}\t{total_crank-ok_Crank-ng_Crank}\nHead\t{total_head}\t{ok_Head}\t{ng_Head}\t{total_head-ok_Head-ng_Head}\n\nDaily TBM card raise status is attached.\nThank You\n\nMNT-MES", 'plain')
my_message = tabular_table_OM.get_html_string()
my_message2 = tabular_table.get_html_string()
html = """\
<html>
    <head>
    <style>
        table, th, td {
            border: 1px solid black;
            border-collapse: collapse;
        }
        th, td {
            padding: 5px;
            text-align: left;    
        }    
    </style>
    </head>
<body>
<p>
Good Morning<br> 
SUMMARY OF PREVIOUS DAY <b>(%s)</b> OM CHECK ACTIVITY<br> 
</p>
<p>    
<b>TNGA-PRODUCTION</b>
    <br>
    %s
</p>
<p>
%s<br>
Thank You
</p>
<p>
MNT-MES
</p>
</body>
</html>
""" % (yesterday_date, my_message, card_message)

textPart1 = MIMEText(html, 'html')


# Creation of a MIMEApplication Part

if card["success"] == True:
    filename = "card_status_tnga.xlsx"
    filePart = MIMEApplication(open(filename, "rb").read(), Name=filename)
    filePart["Content-Disposition"] = 'attachment; filename="%s' % filename

    # Parts attachment
    message.attach(filePart)
message.attach(textPart1)

# Send Email and close connection
server.send_message(message)
server.quit()

# ==================================EMAIL PART END===============================================
